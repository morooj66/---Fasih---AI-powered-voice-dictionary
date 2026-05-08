import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
LEXICON_PATH = BASE_DIR / "data" / "lexicon.json"
DATASET_PATH = Path(r"C:\Users\alsuh\Downloads\AI Glossary - Dataset.xlsx")


def clean_text(value):
    return str(value or "").strip()


def glossary_entry(arabic_term, arabic_definition, english_term):
    return {
        "word": arabic_term,
        "root": "غير متوفر",
        "meanings": [
            {
                "meaning": arabic_definition,
                "example": f"المصطلح الإنجليزي: {english_term}.",
                "synonyms": [],
                "antonyms": [],
                "source": "AI Glossary - Dataset"
            }
        ]
    }


def main():
    with LEXICON_PATH.open("r", encoding="utf-8") as file:
        lexicon = json.load(file)

    workbook = load_workbook(DATASET_PATH, read_only=True, data_only=True)
    sheet = workbook["English - Arabic"]

    imported_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        english_term, _english_definition, arabic_term, arabic_definition = row[:4]
        english_term = clean_text(english_term)
        arabic_term = clean_text(arabic_term)
        arabic_definition = clean_text(arabic_definition)

        if not arabic_term or not arabic_definition:
            continue

        lexicon[arabic_term] = glossary_entry(arabic_term, arabic_definition, english_term)
        imported_count += 1

    with LEXICON_PATH.open("w", encoding="utf-8") as file:
        json.dump(lexicon, file, ensure_ascii=False, indent=2)
        file.write("\n")

    print(f"Imported {imported_count} glossary entries into {LEXICON_PATH}")


if __name__ == "__main__":
    main()
